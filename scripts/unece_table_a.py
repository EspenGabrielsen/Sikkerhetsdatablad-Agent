"""
Extract "Table A: Dangerous Goods List" from the UNECE ADR 2025 PDF.

The table spans pages 300-551 (0-indexed: 299-550).
Each spread (2 pages) contains the left half (cols 1-14) and right half (cols 15-20).
These must be concatenated horizontally to form complete rows.
"""

import csv
import re
from pathlib import Path

import pdfplumber

# ── Configuration ──────────────────────────────────────────────────────────
PDF_PATH = Path(
    r"C:\Users\E1732506\OneDrive - Saint-Gobain\PycharmProjects\Markdowns\2412006_E_ECE_TRANS_352_Vol.I_WEB_0.pdf"
)
OUTPUT_CSV = Path(__file__).parent / "table_a_dangerous_goods.csv"
OUTPUT_XLSX = Path(__file__).parent / "table_a_dangerous_goods.xlsx"

# Page range (0-indexed)
FIRST_PAGE = 299  # page 300 in PDF (first data page)
LAST_PAGE = 550  # page 551 in PDF (last data page)

# Column headers (20 unique columns)
COLUMN_NAMES = [
    "UN_No",  # (1)
    "Name_and_description",  # (2)
    "Class",  # (3a)
    "Classification_code",  # (3b)
    "Packing_group",  # (4)
    "Labels",  # (5)
    "Special_provisions",  # (6)
    "Limited_quantities",  # (7a)
    "Excepted_quantities",  # (7b)
    "Packaging_instructions",  # (8)
    "Special_packing_provisions",  # (9a)
    "Mixed_packing_provisions",  # (9b)
    "Portable_tank_instructions",  # (10)
    "Portable_tank_special_prov",  # (11)
    "ADR_tank_code",  # (12)
    "Tank_special_provisions",  # (13)
    "Vehicle_tank_code",  # (14)
    "Transport_category",  # (15)
    "Tunnel_restriction_code",  # (16)
    "Hazard_identification_No",  # (19) — skip duplicate UN_No (20)
]

# Known column counts per side
LEFT_COLS = 14  # pages 300, 302, 304... (even pages in spread)
RIGHT_COLS = 11  # pages 301, 303, 305... (odd pages in spread)


def extract_cell_text(page, bbox_str):
    """Extract text from a cell given its bounding box string like '(x0, y0, x1, y1)'."""
    if bbox_str is None or bbox_str == "None":
        return ""
    try:
        # Parse "(x0, y0, x1, y1)"
        coords = tuple(float(x) for x in bbox_str.strip("()").split(","))
        cropped = page.within_bbox(coords)
        text = cropped.extract_text() or ""
        return text.strip()
    except Exception:
        return ""


def is_header_row(cells_text):
    """Check if a row is a header row (contains column numbers like '(1)' or '(3a)')."""
    combined = " ".join(cells_text)
    return bool(re.search(r"\(\d+[a-z]?\)", combined))


def is_separator_row(cells_text):
    """Check if a row is a separator (all cells empty or just dashes)."""
    combined = " ".join(cells_text).strip()
    return combined == "" or all(c in "–-" for c in combined.replace(" ", ""))


def extract_table_from_page(page, expected_cols):
    """
    Extract data rows from a single page's main table.
    Returns list of lists (each inner list = one row of cell texts).
    """
    tables = page.find_tables()
    if not tables:
        return []

    # The main table is the first one with many rows
    main_table = None
    for tbl in tables:
        if len(tbl.rows) > 10 and len(tbl.columns) >= expected_cols - 2:
            main_table = tbl
            break

    if main_table is None:
        return []

    rows = []
    for row in main_table.rows:
        cells = [extract_cell_text(page, str(c)) for c in row.cells]

        # Skip header rows and separators
        if is_header_row(cells) or is_separator_row(cells):
            continue

        # Pad/truncate to expected column count
        while len(cells) < expected_cols:
            cells.append("")
        cells = cells[:expected_cols]

        rows.append(cells)

    return rows


def merge_left_right(left_rows, right_rows):
    """
    Merge left (cols 1-14) and right (cols 15-20) rows horizontally.
    Right side has cols (15)-(20) but also repeats (1)-(2) — we skip the duplicates.
    Right side columns: [col15, col16, col17, col18, col19, col20, col1_dup, col2_dup]
    Actually from inspection: right table has 11 columns:
      [0]=col15, [1]=col16, [2]=col17, [3]=col18, [4]=col19, [5]=col20,
      [6]=col1_dup, [7]=col2_dup, [8]=col15_dup?, [9]=col16_dup?, [10]=col17_dup?
    We take the first 6 columns from right side (cols 15-20).
    """
    merged = []
    max_rows = min(len(left_rows), len(right_rows))

    for i in range(max_rows):
        left = left_rows[i]
        right = right_rows[i]

        # Take first 6 columns from right side (cols 15-20)
        right_cols = right[:6] if len(right) >= 6 else right + [""] * (6 - len(right))

        merged_row = left + right_cols
        merged.append(merged_row)

    return merged


def clean_cell(text):
    """Clean up cell text: normalize whitespace, remove stray newlines."""
    text = re.sub(r"\s+", " ", text)
    return text.strip()


def main():
    print(f"Opening PDF: {PDF_PATH}")
    pdf = pdfplumber.open(PDF_PATH)
    print(f"Total pages: {len(pdf.pages)}")

    all_rows = []
    total_data_rows = 0

    # Process each spread (2 pages = left + right)
    for page_idx in range(FIRST_PAGE, LAST_PAGE, 2):
        left_page = pdf.pages[page_idx]
        right_page = pdf.pages[page_idx + 1]

        left_rows = extract_table_from_page(left_page, LEFT_COLS)
        right_rows = extract_table_from_page(right_page, RIGHT_COLS)

        if not left_rows or not right_rows:
            print(
                f"  ⚠ Page {page_idx+1}-{page_idx+2}: No data found (left={len(left_rows)}, right={len(right_rows)})"
            )
            continue

        merged = merge_left_right(left_rows, right_rows)
        all_rows.extend(merged)
        total_data_rows += len(merged)

        # Progress
        if (page_idx - FIRST_PAGE) % 20 == 0:
            print(
                f"  Processed pages {page_idx+1}-{page_idx+2} ({len(merged)} rows, total: {total_data_rows})"
            )

    print(f"\nTotal data rows extracted: {total_data_rows}")

    # Clean all cells
    cleaned = []
    for row in all_rows:
        cleaned.append([clean_cell(c) for c in row])

    # Write CSV
    print(f"\nWriting CSV: {OUTPUT_CSV}")
    with open(OUTPUT_CSV, "w", newline="", encoding="utf-8-sig") as f:
        writer = csv.writer(f)
        writer.writerow(COLUMN_NAMES)
        writer.writerows(cleaned)
    print(f"  → {len(cleaned)} rows written")

    # Write XLSX (if openpyxl available)
    try:
        import openpyxl
        from openpyxl.utils import get_column_letter

        print(f"\nWriting XLSX: {OUTPUT_XLSX}")
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Table A - Dangerous Goods"

        # Header row
        for ci, name in enumerate(COLUMN_NAMES, 1):
            cell = ws.cell(row=1, column=ci, value=name)
            cell.font = openpyxl.styles.Font(bold=True)

        # Data rows
        for ri, row in enumerate(cleaned, 2):
            for ci, val in enumerate(row, 1):
                ws.cell(row=ri, column=ci, value=val)

        # Auto-width
        for ci in range(1, len(COLUMN_NAMES) + 1):
            max_len = len(COLUMN_NAMES[ci - 1])
            for ri in range(2, min(len(cleaned) + 2, 100)):  # sample first 100 rows
                val = ws.cell(row=ri, column=ci).value
                if val:
                    max_len = max(max_len, len(str(val)))
            ws.column_dimensions[get_column_letter(ci)].width = min(max_len + 2, 60)

        wb.save(OUTPUT_XLSX)
        print(f"  → {len(cleaned)} rows written")
    except ImportError:
        print("  (openpyxl not available, skipping XLSX)")

    print("\n✅ Done!")
    return cleaned


if __name__ == "__main__":
    rows = main()

    # Show first 3 rows as preview
    print("\n── Preview (first 3 rows) ──")
    for i, row in enumerate(rows[:3]):
        print(f"\nRow {i+1}:")
        for ci, (name, val) in enumerate(zip(COLUMN_NAMES, row)):
            if val:
                print(f"  {name}: {val[:80]}")
